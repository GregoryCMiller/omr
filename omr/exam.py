#!/usr/bin/python
#Copyright (C) 2013 Greg Miller <gmill002@gmail.com>
"""
==================================
Bubble Vision: Optical Mark Reader
==================================

Extract answer choices from scanned jpg bubble forms.

Graphical User Interface
------------------------
::
    
    $ exam.py


Command Line
------------
::
    
    $ exam.py imagedir [options]


imagedir           
  Input image directory (front side). Lowest numbered image identifies the key.

`--backdir=BACKDIR`
  Optional back side image directory                                                

`--form=FORM`        
  Set the form string (default and only supported="882E")                       

`--help`             
  Show this help message and exit                                               


Output
------

validation images
    Answer bubble means and reference box fits drawn over each input
    image.
    
results.xlsx
    summary            
        Image path, name box image, and total score for each test.
    
    questioninfo       
        Answer choice counts by question. Key excluded.
    
    scoring            
        Answer choice matches key (0/1). Same indices as choices. Score
        is 0 if key is -1.
    
    choices            
        Answer choice matrix. Tests in rows and questions in columns.
        0-4=A-E, -1=n/a.


Install
-------
::
    
    $ pip install omr
    $ pip install --upgrade omr
    $ pip uninstall omr
    
* Requirements

  * `python <http://www.python.org>`_ 2.7+
  * `pip <http://www.pip-installer.org/en/latest/installing.html>`_ ``$ easy_install pip``

* Dependencies (installed by pip)

  * `numpy 1.8.0 <http://www.numpy.org>`_ numerical array object
  * `pillow 2.2.1 <http://python-imaging.github.io/>`_ image module
  * `openpyxl 1.6.2 <http://openpyxl.readthedocs.org/en/latest/>`_ write xlsx files
  
"""
import argparse
import glob
import itertools
import logging
import multiprocessing 
import os
from os.path import abspath, basename, dirname, join
import re
import shutil
import string
import subprocess
import sys
import Tkinter
from tkFileDialog import askdirectory

import numpy as num
from PIL import Image
import openpyxl, openpyxl.drawing

LOG = multiprocessing.get_logger()

def Main(frontdir, form, backdir=None, pool=None):
    """Main application
    
    - process front side
    - (optional) process and join back side 
    - write output
    """
    images, choices = exam_group(frontdir, FORMS[form]['front'], pool)
    
    if backdir:
        backimages, backchoices = exam_group(backdir, FORMS[form]['back'], pool)
        choices = num.hstack((choices, backchoices))
            
    write_exam_group(images, choices, join(frontdir,'OMR'))

    LOG.info('Completed')

def exam_group(testdir, form, pool=None, outname='OMR'):
    """process all tests in a directory returning image path list and 
    choice matrix
    
    - find .jpg images, sort in place by first 2 numeric blocks
    - Create output directory in input test image dir.  
    - Run each test (possibly in parallel). 
    """
    images = glob.glob(join(abspath(testdir), '*.jpg'))
    images.sort(key=lambda x: float(".".join(re.findall('[0-9]+', basename(x))[:2])))
    if not images:
        raise StandardError('at least one image is required')
    
    outdir = join(abspath(testdir), outname)
    shutil.rmtree(outdir, ignore_errors=True)
    [os.makedirs(join(outdir, s)) for s in ['', 'validation', 'names']]
    
    args = [[im, form()] for im in images]
    
    if pool:
        choice_list = pool.map(star_process_exam, args, chunksize=1)
    else:
        choice_list = map(star_process_exam, args)
        
    return images, num.vstack(choice_list)

def write_exam_group(images, choices, outdir, csv=False, xls=True):
    """write exam group output
    
    - Score tests using first image as the key.  
    - Count choice frequency by question. 
    - Write csv and xlsx data files 
    """
    key = num.array(choices[0])   # key is the first test
    key[key == -1] = -2           # -2 key allows -1 tests to score 0
    scoring = choices == key      # score the tests
    total_score = num.sum(scoring, axis=1)
    
    counts_header = ['Question', 'Key', 'CorrectCount', 
                     'None(-1)', 'A(0)', 'B(1)', 'C(2)', 'D(3)', 'E(4)']
                     
    counts = num.zeros((choices.shape[1], 9))
    counts[:, 0] = 1 + num.arange(choices.shape[1])    # question
    counts[:, 1] = choices[0, :]                       # key
    counts[:, 2] = num.sum(scoring[1:, :], axis=0)     # correct count by question
    for i in range(counts.shape[0]):                   # choice frequencies by question
        counts[i, 3:9], x = num.histogram(choices[1:, i], bins=range(-1, 6))
    
    if csv:
        num.savetxt(join(outdir, 'imagefiles.csv'), map(basename, images), fmt='%s')     
        num.savetxt(join(outdir, 'choices.csv'), choices, fmt='%i', delimiter=',')        
        num.savetxt(join(outdir, 'scoring.csv'), scoring, fmt='%i', delimiter=',')
        num.savetxt(join(outdir, 'questioninfo.csv'), counts, fmt='%i', delimiter=',', 
                    header=",".join(counts_header))

    if xls:
        name_files = glob.glob(join(outdir, 'names', "*"))
        wb = openpyxl.Workbook()
        wb = write_xls_images(wb, name_files, total_score, 'summary', header=['Info', 'Score', 'File'], 
                              height=23, width=[47, 5, 20], scale=[0.65, 0.65])
                              
        wb = write_xls_array(wb, counts, 'question info', counts_header, width=6)
        wb = write_xls_array(wb, scoring.astype('i'), 'scoring', width=3)
        wb = write_xls_array(wb, choices, 'choices', width=3)    
        wb.save(join(outdir, 'results.xlsx'))
    
    
def star_process_exam(args):
    """wrapper for process_exam() that takes list of args"""
    return process_exam(*args)
    
def process_exam(imfile, form):
    """Process input test image returning answer choices
    
    - load image (load, check dpi, trim margins, check size)      
    - fit image reference boxes
    - extract answer bubble means, choices
    - write name image 
    - write validation image 
    
    """
    # load image (load, check dpi, trim margins, check size)      
    img = form.load_image(imfile)
    img = form.trim_margins(img)     
    form.check_size(img)
    
    # fit image reference boxes
    if form.refzone:
        meanfit, fit = form.fit_reference(img)    
        img = form.overlay_ref_fit(img, meanfit, fit)    
        form.set_offset(*meanfit)
    
    # extract answer bubble means, choices
    bubble_means = form.get_bubble_means(img)
    choice = form.choose_answers(bubble_means)
    img = form.overlay_bubble_means(img, bubble_means)
    
    # write name image
    name_file = join(dirname(imfile), 'OMR', 'names', basename(imfile)[:-3] + 'png')
    form.write_info_image(img, name_file)
    
    # write validation image 
    val_file = join(dirname(imfile), 'OMR', 'validation', basename(imfile))        
    form.write_validation(img, val_file)
    
    # write status to console 
    LOG.setLevel(logging.INFO)
    LOG.info(basename(imfile))
    LOG.setLevel(logging.WARN)
    
    return choice

class Form:
    """Represents a generic exam form. Specify answer grid size properties, reference and info rectangles 

    
    Form specification
    ------------------  
    
    Grid parameters specified as [height, width] in pixels.  
    
    ================  ===============================================================================
    Parameter         Description
    ================  ===============================================================================
    size              h,w size of answer bubble matrix (n questions, n answer choices)
    pos               h,w coordinate of answer matrix upper left corner (i0, j0) in pixels
    bub               h,w answer bubble surrounding box in pixels (float ok)         
    space             h,w unit cell edge lengths in pixels (float ok)
    offset            h,w fitted reference offset applied to pos
    ================  ===============================================================================
    
    ::
        
        |
        |        |-bub-|
        |        |---space---|
        |
        |       j0    j1    j2
        |        ___________________
        |    i0 |0                  |  
        |       | [ A ]       [ B ] |  
        |    i1 |      1            |  
        |       |                   |
        |    i2 |            2      |
        |       | [ A ]       [ B ] |
        |       |___________________|
        |


    Region rectangles
    -----------------

    rectangels are specified as [height min, height max, width min, width max]
    
    ================  ==========================================================
    Parameter         Description
    ================  ==========================================================
    refzone           list of black reference box rectangles (or empty list). boxes go 
    info              name info rectangle (or None)
    score             machine printed score rectangle (or None)    
    ================  ==========================================================
    
    Image properties 
    ----------------
    
    ================  ====================================================================
    Parameter         Description
    ================  ====================================================================
    expected_dpi      h,w image dpi (CRITICAL - relates pixels to distance)
    expected_size     h,w expected image size (after conversion to proper dpi) 
    size_tolerance    allowed percent error in actual image size 
    contrast          black/white contrast split value 0<=x<=255
    trim_std          minimum stdev to remove image edge during trimming
    radius            reference box fitting search radius in pixels
    min_ref           minimum pixel value for black box match 0<=x<=255
    ref_x, ref_y      validation image reference fit summary panel coordinates
    signal            minimum ratio of darkest to second darkest answer choice    
    ================  ====================================================================
    
    
    """
    size           = [0, 0]  
    offset         = [0, 0]
    pos            = [0, 0] 
    bub            = [0, 0]        
    space          = [0, 0]
    
    info           = None # [0, 0, 0, 0]
    score          = None # [0, 0, 0, 0]   
    refzone        = None 
                     # [[0, 0, 0, 0],
                      #[0, 0, 0, 0], 
                      #[0, 0, 0, 0],               
                      #[0, 0, 0, 0]]
    
    expected_dpi   = [0, 0]
    expected_size  = [0, 0] 
    size_tolerance = [0, 0]
    ref_x, ref_y   = 0, 0 
    contrast       = 0.0 * 255 
    trim_std       = 0         
    radius         = 0         
    min_ref        = 0.0 * 255 
    signal         = 0.0       
    
    def __init__(self):
        """initialize form, calculate default coordinates.  """
        self.calc_coords()

    def calc_coords(self):
        """calculate (m, n, 4) sized matrix of answer bubble
        hmin,hmax,wmin,wmax coordinates"""
        i = num.outer(num.arange(self.size[0]), num.ones(self.size[1]))
        i0 = self.pos[0] + (i * self.space[0])
        i1 = self.pos[0] + (i * self.space[0]) + self.bub[0]
        
        j = num.outer(num.ones(self.size[0]), num.arange(self.size[1]))        
        j0 = self.pos[1] + (j * self.space[1])
        j1 = self.pos[1] + (j * self.space[1]) + self.bub[1]
        
        self.coords = num.dstack((i0,i1,j0,j1)).astype('i')
    
    def set_offset(self, r=0, c=0):
        """update positional parameters with offset, recalculate
        coordinates matrix"""
        self.offset = num.array(self.offset) + num.array([r, c])
        self.pos = [self.pos[0] + r, self.pos[1] + c]

        if self.info:
            self.info = num.array(self.info) + num.array([r, r, c, c])
    
        if self.score:
            self.score = num.array(self.score) + num.array([r, r, c, c])
        
        self.calc_coords()
    
    def load_image(self, imfile):
        """open input image, correct dpi, return greyscale array"""
        im = Image.open(imfile)          # im = open as PIL image 
        
        dpi_ratio = num.true_divide(self.expected_dpi, num.array(im.info['dpi']))
        newsize = (num.array(im.size) * dpi_ratio).astype('i')
        if not all(newsize == num.array(im.size)):
            im = im.resize(newsize, Image.BICUBIC) # change dpi
        
        img = num.array(im.convert('L')) # convert to greyscale array 0-255
        
        return img

    def trim_margins(self, img):
        """Recursivly trim blank edges (low stdev) from input array"""
        oldsize = (0, 0)
        while oldsize != img.shape: # while the size is changing
            oldsize = img.shape
            for i in range(4):                         # 4 times
                img = num.rot90(img)                   #   rotate 90
                if num.std(img[0, :]) < self.trim_std: #   if low std
                    img = img[1:, :]                   #     trim edge 
    
        return img
        
    def check_size(self, img):
        """Check input image dimensions are within form tolerance. """
        absdiff = num.abs(num.subtract(img.shape, self.expected_size))
        pctdiff = num.true_divide(absdiff, self.expected_size)
        if not num.all(pctdiff <= self.size_tolerance):
            raise StandardError('image size outside form tolerance {} != {}'\
                                    .format(img.shape, self.expected_size))
        
    def fit_reference(self, img):
        """Get the best translation offset by fitting black box
        reference zones"""
        bw_img = 255 * (img >= self.contrast) 
        fit = [center_on_box(bw_img, self.radius, self.min_ref, *ref) for ref in self.refzone]
        meanfit = num.mean(num.ma.masked_array(fit, fit == -9999), axis=0).astype('i')
        if meanfit[0] is num.ma.masked:
            raise StandardError('At least one reference box match required')
    
        return meanfit, fit                  

    def get_bubble_means(self, img):
        """get the mean pixel value in each answer bubble region"""
        bw_img = 255 * (img >= self.contrast)
        means = num.zeros(self.coords.shape[:2])     
        for (i,j) in itertools.product(*map(range, self.size)):
            i0, i1, j0, j1 = self.coords[i, j, :]
            means[i, j] = num.mean(bw_img[i0:i1, j0:j1])

        return means
    
    def choose_answers(self, means):     
        """choose darkest answer choice. assign poor signal choices -1"""
        choice = num.argmin(means, axis=1)
        if self.signal:
            sorted_rows = num.sort(means, axis=1)
            signal = sorted_rows[:,1] / sorted_rows[:,0]
            choice[signal <= self.signal] = -1
            
        return choice

    def overlay_ref_fit(self, img, mean, fit, off=25):
        """draw crosses at the corners of the initial and fitted
        reference boxes"""
        def plus(img, x, y, val=0, r=10):
            img[x-1:x, y-r:y+r], img[x-r:x+r, y-1:y] = val, val
            return img
        
        centers = [(self.ref_x - off, self.ref_y - off), 
                   (self.ref_x - off, self.ref_y + off), 
                   (self.ref_x + off, self.ref_y - off), 
                   (self.ref_x + off, self.ref_y + off)]
                           
        img = plus(img, self.ref_x, self.ref_y, val=150, r=15) # final mean offset
        img = plus(img, self.ref_x + mean[0], self.ref_y + mean[1], val=0)        
        for [x0,x1,y0,y1], [x_off, y_off], (cx, cy) in zip(self.refzone, fit, centers):
            img = plus(img, cx,  cy, val=120, r=15)        # panel fitted
            img = plus(img, cx + x_off, cy + y_off, val=0) # panel reference            
            img = plus(img, x0, y0, val=150)               # expected reference
            img = plus(img, x1, y1, val=150)               #
            img = plus(img, x0 + x_off, y0 + y_off, val=0) # actual fitted
            img = plus(img, x1 + x_off, y1 + y_off, val=0) # 
        
        return img
    
    def overlay_bubble_means(self, img, means):
        """overlay the bubble region mean values onto the image"""
        for (i,j) in itertools.product(*map(range, self.size)):
            i0, i1, j0, j1 = self.coords[i, j, :]
            img[i0:i1, j0:j1] = means[i,j]
    
        return img
    
    def write_validation(self, img, val_file):
        """write the greyscale image that has been marked with reference
        fits and bubble means"""
        Image.fromarray(img).save(val_file)

    def write_info_image(self, img, name_file):
        """extract the forms info box region and stack the score box"""
        if self.info not in [[], None]:
            xmin, xmax, ymin, ymax = self.info
            nameimg = num.rot90(img[xmin:xmax, ymin:ymax])
    
            if self.score not in [[], None]:
                xmin, xmax, ymin, ymax = self.score
                score = num.rot90(img[xmin:xmax, ymin:ymax])
                nameimg = num.hstack([nameimg[30:75, :], score])

            Image.fromarray(nameimg).save(name_file)


class Form882E_front(Form):
    """scantron form 882E front side or equivilant"""
    size    = [50, 5]
    pos     = [258, 130]
    space   = [25.2, 49.2]
    bub     = [15, 39]
    info    = [746, 1234, 408, 575]
    score   = [1350, 1395, 360, 405]
    refzone = [[233, 249, 51, 81],
               [106, 125, 571, 601],
               [1574, 1592, 570, 600],
               [1492, 1502, 50, 79]]
    
    expected_dpi = [150, 150]
    expected_size = [1664, 664]
    size_tolerance = [0.04, 0.04]
    ref_y, ref_x = 525, 175
    
    contrast = 0.7 * 255 
    trim_std = 4         
    radius   = 10        
    min_ref  = 0.5 * 255 
    signal   = 1.10      
    
class Form882E_back(Form882E_front):
    """scantron form 882E back side or equivilant"""
    info = None
    

# dictionary containing all forms {name:[front,back], }
FORMS = {'882E': {'front':Form882E_front, 'back':Form882E_back}, }
    
def center_on_box(img, radius, min_ref, xmin, xmax, ymin, ymax, na_val=-9999):
    """find the best offset for a black box by trying all within a
    circular search radius"""
    x, y = num.meshgrid(num.arange(-radius, radius), num.arange(-radius, radius))
    coords = [(i, j) for i, j in zip(x.flatten(), y.flatten()) if (i**2 + j**2)**0.5 <= radius]    
    fit = [num.mean(img[(xmin+i):(xmax+i), (ymin+j):(ymax+j)]) for i, j in coords]
    if num.nanmin(fit) <= min_ref:
        return num.array(coords[num.nanargmin(fit)])
    else:
        return num.array([na_val, na_val]) 

def write_xls_array(workbook, inarray, title=None, header=None, row=0, col=0, width=None, height=None):
    """create a new sheet in xlsx workbook, write header and array. """
    ws = workbook.create_sheet()
    
    if title:
        ws.title = title    
    
    if header:
        [setattr(ws.cell(row=row, column=col+j), 'value', h) for j, h in enumerate(header)]
        row += 1
            
    for i,j in itertools.product(*map(range, inarray.shape)):
        setattr(ws.cell(row=row+i, column=col+j), 'value', inarray[i, j])
    
    if width:
        if not hasattr(width, '__iter__'):
            width = itertools.repeat(width)
    
        for w, a in zip(width, sorted(ws.column_dimensions.keys())):
            setattr(ws.column_dimensions[a], 'width', w)
    
    if height:
        for k in ws.row_dimensions.keys():
            setattr(ws.row_dimensions[k], 'height', height)
    
    return workbook

def write_xls_images(wb, name_images, scores, title=None, header=None, 
    width=None, height=None, scale=[1, 1]):
    """write xlsx file containing a table of extracted info box images,
    score, and file name for each test"""
    ws = wb.get_active_sheet()
    if title:
        ws.title = title
    
    if header:
        [setattr(ws.cell(row=0, column=j), 'value', h) for j, h in enumerate(header)]
        
    for row, (score, name_file) in enumerate(zip(scores, name_images)):
        ws.cell(row=row+1, column=1).value = score
        ws.cell(row=row+1, column=2).value = basename(name_file)
    
    if width:
        for w, a in zip(width, sorted(ws.column_dimensions.keys())):
            setattr(ws.column_dimensions[a], 'width', w)
    
    if height:
        for k in ws.row_dimensions.keys():
            setattr(ws.row_dimensions[k], 'height', height)
    
    if name_images:
        try: 
            size = openpyxl.drawing.Image(name_images[0]).image.size * num.array(scale)
            for r, im in enumerate(name_images):
                img = openpyxl.drawing.Image(im, size=size)
                img.anchor(ws.cell(row=r+1, column=0))
                ws.add_image(img)
        except:
            setattr(ws.cell(row=1, column=0), 'value', 'ERROR: Info images could not be loaded')
    else:
        setattr(ws.cell(row=1, column=0), 'value', 'ERROR: Info images not found')
    
    return wb


class OmrGui(Tkinter.Frame):
    """GUI to select input args"""
    def __init__(self, master):
        """initialize frame, create output variables and widgets, verify
        command and get help text"""
        Tkinter.Frame.__init__(self, master, padx=5, pady=5)
        master.wm_title("Bubble Vision:  Optical Mark Reader")
        self.pack()
        
        self.front = Tkinter.StringVar(self)
        Tkinter.Label(self, text="Front Directory").grid(row=0, column=0)
        Tkinter.Button(self, textvar=self.front, command=self.get_front, width=30).grid(row=1, column=0)
        
        self.back = Tkinter.StringVar(self)
        Tkinter.Label(self, text="Back Directory (optional)").grid(row=0, column=1)
        Tkinter.Button(self, textvar=self.back, command=self.get_back, width=30).grid(row=1, column=1)
        
        self.form = Tkinter.StringVar(self)
        Tkinter.Label(self, text='Form').grid(row=0, column=2)
        Tkinter.OptionMenu(self, self.form, *FORMS.keys()).grid(row=1, column=2)
        self.form.set('882E')
        
        self.text = Tkinter.Text(self, height=15, width=80)
        self.text.grid(row=2, columnspan=3, pady=5)
        
        Tkinter.Button(self, text='Quit', command=self.quit, width=8).grid(row=3, column=0, sticky=Tkinter.W)
        Tkinter.Button(self, text='Run', command=self.run_app, width=8).grid(row=3, column=2, sticky=Tkinter.E)
        
        self.cmd = ['python', abspath(__file__)]
        try:
            self.run_command(['--help'], see="1.0")
        except:
            self.cmd = []
            self.run_app()
    
    def run_app(self):
        """collect and verify user selected arguments, run main
        application"""
        if not self.cmd:
            self.insert_text('ERROR: No application!\n')
            return None
            
        if not self.front.get():
            self.insert_text('ERROR: Choose front directory!\n')
            return None

        args = [self.front.get(), ] 
        args.append('--form={}'.format(self.form.get()))
        if self.back.get():
            args.append('--backdir={}'.format(self.back.get()))
        
        self.run_command(args)
        
    def run_command(self, args, echo=True, output=True, see=Tkinter.END):
        """run the command with input args. Use echo and output to print
        the command and response""" 
        if echo:
            self.insert_text("$ " + " ".join(self.cmd + args) + '\n')
        
        p = subprocess.Popen(self.cmd + args, shell='win' in sys.platform, 
                             stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        
        if output:
            [self.insert_text(l, see) for l in iter(p.stdout.readline,'')]
            
    def insert_text(self, text, see=Tkinter.END):
        """insert string into gui text box"""
        self.text.config(state=Tkinter.NORMAL)
        self.text.insert(Tkinter.END, text)
        self.text.config(state=Tkinter.DISABLED)
        if see:
            self.text.see(see)

        self.text.update_idletasks()
        
    def get_front(self):
        """open directory selection dialog to set front directory or enpty string"""
        self.front.set(askdirectory())

    def get_back(self):
        """open directory selection dialog to set back directory or enpty string"""
        self.back.set(askdirectory())

def command_args():
    """parse command line arguments returning namespace object. adds --help option"""
    parser = argparse.ArgumentParser(description=
        "Extract answer choices from scanned jpg bubble forms. ")
    
    parser.add_argument('frontdir', 
                        help="image directory (front). ")
    
    parser.add_argument('-b', '--backdir', default=None, 
                        help='optional back side image directory')
        
    parser.add_argument('-f', '--form', default='882E', choices=FORMS.keys(), 
                        help='set form')
    
    return parser.parse_args()


if __name__ == '__main__':
    
    if len(sys.argv) > 1:               # has input args: run as command line app
        multiprocessing.log_to_stderr()     
        args = command_args()              
        args.pool = multiprocessing.Pool()  
        LOG.setLevel(logging.INFO)          
        
        Main(**vars(args))              # pass input as kwargs to main app

        LOG.setLevel(logging.WARN)            
        args.pool.close()                   
        args.pool.join()

    else:                               # no input args: run as GUI app
        root = Tkinter.Tk()
        app = OmrGui(root)
        root.update_idletasks()   
        root.mainloop()
        