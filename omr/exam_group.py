#Copyright (C) 2013 Greg Miller <gmill002@gmail.com>
"""process a group of test images contained in a directory"""

from functools import partial
from glob import glob
from itertools import repeat, product
from os import mkdir
from os.path import basename, join
from re import findall
from shutil import rmtree
from numpy import array, histogram, hstack, savetxt, sum, vstack, zeros

try:
    import openpyxl
    import openpyxl.drawing
except ImportError:
    openpyxl = None

from omr import FORMS, process_exam

_NUMSORT = lambda x: float(".".join(findall('[0-9]+', basename(x))[:2]))
"""extract the first two numeric blocks of a path as a float"""


def main(frontdir, form, backdir=None, pool=None):
    """Main command line application. """

    fimg, fchoice, fout = process_exam_group(frontdir, form, 'front', pool)

    if backdir:
        bimg, bchoice, bout = process_exam_group(backdir, form, 'back', pool)
        fchoice = hstack((fchoice, bchoice))

    write_exam_group(fimg, fchoice, fout)


def process_exam_group(testdir, formstr, side, pool=None):
    """Process all test images in a directory returning image path list and 
    choice matrix. 
    
    
    testdir  
        Input folder containing test images 
    
    formstr  
        String identifying form in FORMS dictionary

    side     
        Test side (front, back)

    pool     
        Parallel processing pool 

    
    Procedure
    
    - Create output directory in input test image dir.
    - find .jpg images, sort in place by first 2 numeric blocks
    - Run each test (possibly in parallel). 
        
    
    """
    # define output directories 
    wd = join(testdir, 'OMR')
    rmtree(wd, True)
    [mkdir(join(wd, p)) for p in ['', 'validation', 'names']]

    # get image paths
    images = sorted(glob(join(testdir, '*.jpg')), key=_NUMSORT)
    if not images:
        raise StandardError('at least one image is required')

    # process each image
    func = partial(process_exam, formcfg=FORMS[formstr][side])
    if pool:
        choices = pool.map(func, images)
    else:
        choices = map(func, images)

    # return image list, choices, and output direcory
    return images, vstack(choices), wd


def write_exam_group(images, choices, outdir):
    """Write exam group output
    
    
    images   
        list of image paths 
    
    choices  
        matrix of choices with tests in rows
    
    outdir   
        output path

    - Score tests using first image as the key.  
    - Count choice frequency by question. 
    - Write csv and xlsx data files 
    
    """
    #score tests
    key = choices[0]  # key is the first test
    key[key == -1] = -2  # -2 key allows -1 tests to score 0
    scoring = choices == key  # score all
    score_by_test = sum(scoring, 1)  # score by test
    score_by_question = sum(scoring[1:, :], 0)  # exclude key

    # count choice frequency
    counts = zeros((choices.shape[1], 9))
    counts[:, 0] = range(1, 1 + choices.shape[1])  # question number
    counts[:, 1] = choices[0, :]  # correct choice
    counts[:, 2] = score_by_question  # correct count by question (ex key)
    for i in range(counts.shape[0]):  # choice frequencies by question
        counts[i, 3:9], _x = histogram(choices[1:, i], range(-1, 6))

    counts_header = ['Question', 'Key', 'CorrectCount',
                     'None(-1)', 'A(0)', 'B(1)', 'C(2)', 'D(3)', 'E(4)']

    # csv output
    savetxt(join(outdir, 'imagefiles.csv'), images, fmt='%s')
    savetxt(join(outdir, 'choices.csv'),  choices, fmt='%i', delimiter=',')
    savetxt(join(outdir, 'scoring.csv') , scoring, fmt='%i', delimiter=',')
    savetxt(join(outdir, 'questioninfo.csv') , counts, fmt='%i', delimiter=',',
            header=",".join(counts_header))

    # xls output
    if openpyxl is not None:
        name_files = sorted(glob(join(outdir, 'names', '*')), key=_NUMSORT)

        wb = openpyxl.Workbook()
        wb = write_xls_images(wb, name_files, score_by_test, 'summary')
        wb = write_xls_array(wb, counts, 'question info', counts_header, width=6)
        wb = write_xls_array(wb, scoring.astype('i'), 'scoring', width=3)
        wb = write_xls_array(wb, choices, 'choices', width=3)
        wb.save(join(outdir, 'results.xlsx'))


def write_xls_array(workbook, inarray, title=None, header=None, row=0, col=0, width=None, height=None):
    """write input array to a new sheet in input xlsx workbook
    
        
    workbook  
        xlsx workbook returned by openpyxl     
    
    inarray   
        input array 
    
    title     
        name of sheet 
    
    header    
        header as list of strings 
    
    row       
        starting row 
    
    col       
        starting column
    
    width     
        row width
    
    height    
        row height

    """
    ws = workbook.create_sheet()

    if title:
        ws.title = title

    if header:
        [setattr(ws.cell(row=row, column=col + j), 'value', h) for j, h in enumerate(header)]
        row += 1

    for i, j in product(*map(range, inarray.shape)):
        setattr(ws.cell(row=row + i, column=col + j), 'value', str(inarray[i, j]))

    if width:
        if not hasattr(width, '__iter__'):
            width = repeat(width)

        for w, a in zip(width, sorted(ws.column_dimensions.keys())):
            setattr(ws.column_dimensions[a], 'width', w)

    if height:
        for k in ws.row_dimensions.keys():
            setattr(ws.row_dimensions[k], 'height', height)

    return workbook


def write_xls_images(workbook, name_images, scores, title=None, header=['Info', 'Score', 'File'],
                     height=23, width=[47, 5, 20], scale=[0.65, 0.65]):
    """write xlsx file containing a table of extracted info box images,
    score, and file name for each test"""
    ws = workbook.get_active_sheet()
    if title:
        ws.title = title

    if header:
        [setattr(ws.cell(row=0, column=j), 'value', h) for j, h in enumerate(header)]

    for row, (score, name_file) in enumerate(zip(scores, name_images)):
        ws.cell(row=row + 1, column=1).value = score
        ws.cell(row=row + 1, column=2).value = basename(name_file)

    if width:
        for w, a in zip(width, sorted(ws.column_dimensions.keys())):
            setattr(ws.column_dimensions[a], 'width', w)

    if height:
        for k in ws.row_dimensions.keys():
            setattr(ws.row_dimensions[k], 'height', height)

    if name_images:
        try:
            size = openpyxl.drawing.Image(str(name_images[0])).image.size * array(scale)
            for r, im in enumerate(name_images):
                img = openpyxl.drawing.Image(str(im), size=size)
                img.anchor(ws.cell(row=r + 1, column=0))
                ws.add_image(img)
        except: # TODO specify exception
            setattr(ws.cell(row=1, column=0), 'value', 'ERROR: Info images could not be loaded')
    else:
        setattr(ws.cell(row=1, column=0), 'value', 'ERROR: Info images not found')

    return workbook
